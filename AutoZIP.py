
# import zipfile
import os
import subprocess
import time
from openpyxl import Workbook
from openpyxl import load_workbook
import PySimpleGUI as sg

sleeptime = 5
flag = True
source_file = None
source_file = sg.PopupGetFile("请选择数据源文件：")
if not source_file:
	flag = False
if flag:
	wb = load_workbook(filename=source_file)
	# grab the active worksheet
	ws = wb.active
	fnamelist = []
	lnamelist = []
	files_pathlist = []
	# rarfilepath = r'C:\Users\vicky.hongmei.lu\Desktop\'
	rarfilepath = None
	rarfilepath = sg.PopupGetFolder('请选择存放压缩包的文件夹：')
	if not rarfilepath:
		flag = False
	# print(rarfilepath)
	if flag:

		error_name = []

		for i in range(2,100):
			if ws['G'+str(i)].value:
				fnamelist.append(ws['G'+str(i)].value.split(' ')[1])
				lnamelist.append(ws['G'+str(i)].value.split(' ')[0])
				files_pathlist.append(ws['Q'+str(i)].value)
			else:
				break


				# print(ws['G'+str(i)].value)
		for i in range(len(fnamelist)):
			# print(fnamelist[i], lnamelist[i], 'T:'+files_pathlist[i][33:]+ '\\')
			fname = fnamelist[i]
			lname = lnamelist[i]
			files_path = 'T:'+files_pathlist[i][33:]+ '\\'
			# files_path = r"C:\bb\cc"

			rarfullname = rarfilepath +'\\'+ lname + ' ' +fname + '.rar'
			rarcmd = r'C:\Program Files\WinRAR\WinRAR.exe'
			rarpasswd = lname+'!'+fname

			filecount = 0
			for current_path, subfolders, filesname in os.walk(files_path):
				for file in filesname:
					# print(os.path.join(current_path, file))
					if ("Offer Letter_" in file) and (".pdf" in file):
						filecount = filecount + 1
						cmd = [rarcmd, 'a', '-ep', '-hp'+rarpasswd, rarfullname, os.path.join(current_path, file)]
						sp = subprocess.Popen(cmd, stderr=subprocess.STDOUT, stdout=subprocess.PIPE)
						time.sleep(sleeptime)
					if ("Employment Contract" in file) and ('.pdf' in file):
						filecount = filecount + 1
						cmd = [rarcmd, 'a', '-ep', '-hp'+rarpasswd, rarfullname, os.path.join(current_path, file)]
						sp = subprocess.Popen(cmd, stderr=subprocess.STDOUT, stdout=subprocess.PIPE)
						time.sleep(sleeptime)
					if ('Cover' in file) and ('.pdf' in file):
						filecount = filecount + 1
						cmd = [rarcmd, 'a', '-ep', '-hp'+rarpasswd, rarfullname, os.path.join(current_path, file)]
						sp = subprocess.Popen(cmd, stderr=subprocess.STDOUT, stdout=subprocess.PIPE)
						time.sleep(sleeptime)
					if ('Checklist' in file) and ('.pdf' in file):
						filecount = filecount + 1
						cmd = [rarcmd, 'a', '-ep', '-hp'+rarpasswd, rarfullname, os.path.join(current_path, file)]
						sp = subprocess.Popen(cmd, stderr=subprocess.STDOUT, stdout=subprocess.PIPE)
						time.sleep(sleeptime)
					if ('form' in file) and ('.xlsx' in file):
						filecount = filecount + 1
						cmd = [rarcmd, 'a', '-ep', '-hp'+rarpasswd, rarfullname, os.path.join(current_path, file)]
						sp = subprocess.Popen(cmd, stderr=subprocess.STDOUT, stdout=subprocess.PIPE)
						time.sleep(sleeptime)
			if filecount == 5:
				filecount = 0
			else:
				error_name.append(lname + fname)



		if len(error_name) != 0:
			error_name_list = ''
			for e in error_name:
				error_name_list = error_name_list + e + '\n'
			sg.PopupError('以下候选人文件确实，请手工处理：\n' + error_name_list)
		else:
			sg.PopupOK("压缩包全部生成成功，请检查！")
				